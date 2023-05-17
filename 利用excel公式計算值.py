from openpyxl import load_workbook
from win32com import client as wc
import os
 
def open_and_save(filename):
    xl = wc.DispatchEx("Excel.Application")
    wb = xl.workbooks.open(os.path.abspath(filename))
    xl.Visible = False
    wb.Save()
    xl.Quit()

# 讀取 Excel 檔案
wb = load_workbook('test.xlsx')
actSheet = wb.active
# 修改值
actSheet['A1'] = 1
# # 儲存檔案
wb.save('result.xlsx')

# 模擬打開動作，讓excel計算值
open_and_save('result.xlsx')
# # 讀取 Excel 檔案
wb = load_workbook('result.xlsx', data_only=True)
actSheet = wb.active
print(actSheet['C1'].value)
