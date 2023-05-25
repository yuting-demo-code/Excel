
from openpyxl import load_workbook
import numpy as np
import matplotlib.pyplot as plt
import numpy as np

# Load your workbook and sheet as you want, for example
wb = load_workbook('GM2_分析.xlsm', data_only=True)
ws = wb['Golden_LSC']

# 取得圖表資料
data_range = ws['G1':'W13']

# 將資料轉換為 NumPy 陣列
Z = np.array([[cell.value for cell in row] for row in data_range])

# 顯示 NumPy 陣列
# print(Z)

# 建立 X, Y 二維陣列
x = np.arange(0, Z.shape[1], 1)
y = np.arange(0, Z.shape[0], 1)
X, Y = np.meshgrid(x, y)

# 設置顏色映射
colormap = 'viridis'  # 可以選擇不同的顏色映射
color_map = plt.get_cmap(colormap)

# 繪製表面圖
fig = plt.figure()
ax = fig.add_subplot(111, projection='3d')
surf = ax.plot_surface(X, Y, Z, cmap=color_map)

plt.show()
